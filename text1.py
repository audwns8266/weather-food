from pyowm import OWM
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5 import QtCore, QtGui, QtWidgets
from datetime import datetime
import threading
import openpyxl
import random
import webbrowser

API_key = "10d4923cb75c31bfa0a786005f4803db" #openweathermap API
owm = OWM(API_key)

datetime.today()        # 시간 가져오기
datetime.today().year
datetime.today().month
datetime.today().day
datetime.today().hour

wb = openpyxl.load_workbook('음식.xlsx') #엑셀 파일 읽기
ws = wb.active # ws = wb.get_sheet_by_name("Sheet1")

class MyWindow(QMainWindow):

    def __init__(self):
        super().__init__()
      
        self.setupUI()
        
       
    def setupUI(self):
        palette = QtGui.QPalette() # 배경색 지정
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        self.setPalette(palette)

        
        font = QtGui.QFont() # 폰트 지정
        font.setFamily("휴먼옛체")
        font.setPointSize(10)
        self.setFont(font)
        
        self.setGeometry(800, 400, 300, 350) # 창 크기 설정
        self.setWindowTitle('W-F') # 창 이름 설정
        self.setWindowIcon(QIcon('아이콘.png')) # 아이콘 설정
        
        textLabel = QLabel("지역 : ",self) # 텍스트 출력
        
        textLabel.move(40, 30) # 텍스트 위치 설정

        self.lineEdit = QLineEdit("",self) #lineEdit 생성
        self.lineEdit.setAlignment(QtCore.Qt.AlignCenter) #텍스트 가운데 정렬
        self.lineEdit.move(100,30) #lineEdit 위치 설정
        self.lineEdit.returnPressed.connect(self.lineEditPressed) #엔터를 누르면 lineEditPressed 실행

        self.label = QLabel("", self) # QLabel 생성
        self.label.setAlignment(QtCore.Qt.AlignCenter) #텍스트 가운데 정렬
        
        self.image = QLabel(self) # QLabel 생성
        self.image.setAlignment(QtCore.Qt.AlignCenter) #텍스트 가운데 정렬

        btn1 = QPushButton("추천 메뉴", self) # 버튼 생성
        btn1.setStyleSheet('QPushButton {background-color: white}')
        btn1.move(40, 280)
        btn1.clicked.connect(self.btn1_clicked)

        self.clock = QLabel("",self) # 시간 Label 생성
        self.clock.move(140,320)
        self.clock.resize(200, 20)
        self.Time()

        btn2 = QPushButton("음식점 검색", self) # 버튼 생성
        btn2.setStyleSheet('QPushButton {background-color: white}')
        btn2.move(160, 280)
        btn2.clicked.connect(self.btn2_clicked)
        
    def Time(self):
        self.clock.setText(datetime.today().strftime("%Y/%m/%d %H:%M:%S"))
        threading.Timer(1,self.Time).start()#1초마다 업데이트        

    def btn1_clicked(self):
        self.image.move(55, 200) # QLabel 위치 설정
        self.image.resize(210, 100) # QLabel 크기 설정
        
        self.label.move(75, 70) # QLabel 위치 설정
        self.label.resize(150, 150) # QLabel 크기 설정
        Rain_B = 'B3','B4','B5','B6','B7','B8','B9','B10','B11' #비오는날 아침
        Rain_R = 'B12','B13','B14','B15','B16','B17','B18','B19','B20', 'B21' #비오는날 점심
        Rain_D = 'B22','B23','B24','B25','B26','B27','B28','B29','B30' #비오는날 저녁

        Snow_B = 'E3','E4','E5','E6','E7','E8','E9','E10','E11' #눈오는날 아침
        Snow_R = 'E12','E13','E14','E15','E16','E17','E18','E19','E20', 'E21' #눈오는날 점심
        Snow_D = 'E22','E23','E24','E25','E26','E27','E28','E29','E30' #눈오는날 저녁

        Clear_B = 'K3','K4','K5','K6','K7','K8','K9','K10','K11' #맑은날 아침
        Clear_R = 'K12','K13','K14','K15','K16','K17','K18','K19','K20', 'K21' #맑은날 점심
        Clear_D = 'K22','K23','K24','K25','K26','K27','K28','K29','K30' #맑은날 저녁

        Clouds_B = 'H3','H4','H5','H6','H7','H8','H9','H10','H11' #흐린날 아침
        Clouds_R = 'H12','H13','H14','H15','H16','H17','H18','H19','H20', 'H21' #흐린날 점심
        Clouds_D = 'H22','H23','H24','H25','H26','H27','H28','H29','H30' #흐린날 저녁

        Mist_B = Rain_B + Snow_B + Clear_B + Clouds_B # 안개낀날은 맑음,흐림,비,눈 중 랜덤
        Mist_R = Rain_B + Snow_B + Clear_B + Clouds_B
        Mist_D = Rain_B + Snow_B + Clear_B + Clouds_B

        Haze_B = Mist_B # Haze와 Mist 모두 안개
        Haze_R = Mist_R
        Haze_D = Mist_D

        Fog_B = Mist_B # Haze와 Mist 모두 안개
        Fog_R = Mist_R
        Fog_D = Mist_D
        
        try :
            city = str(self.lineEdit.text()) # city에 lineEdit에 써진 텍스트 입력 
            obs = owm.weather_at_place(city) #날씨를 검색할 장소를 city로 설정
            w = obs.get_weather() # 날씨 정보 가져오기
            l = obs.get_location() # 위치 정보 가져오기
            

            if w.get_status() == 'Snow' and datetime.today().strftime("%H")<str(12): # 눈오는날 아침
                Snow_B = random.choice(Snow_B)
                food = ws[Snow_B].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Snow' and datetime.today().strftime("%H")<str(18) and datetime.today().strftime("%H")>str(12): #눈오는날 점심
                Snow_R = random.choice(Snow_R)
                food = ws[Snow_R].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Snow' and datetime.today().strftime("%H")<str(24) and datetime.today().strftime("%H")>str(18): #눈오는날 저녁
                Snow_D = random.choice(Snow_D)
                food = ws[Snow_D].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            
            if w.get_status() == 'Rain'and datetime.today().strftime("%H")<str(12): #비오는날 아침
                Rain_B = random.choice(Rain_B)
                food = ws[Rain_B].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Rain'and datetime.today().strftime("%H")<str(18) and datetime.today().strftime("%H")>str(12): #비오는날 점심
                Rain_R = random.choice(Rain_R)
                food = ws[Rain_R].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Rain'and datetime.today().strftime("%H")<str(24) and datetime.today().strftime("%H")>str(18): #비오는날 저녁
                Rain_D = random.choice(Rain_D)
                food = ws[Rain_D].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
        
            if w.get_status() == 'Clear'and datetime.today().strftime("%H")<str(12): # 맑은날 아침
                Clear_B = random.choice(Clear_B)
                food = ws[Clear_B].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Clear'and datetime.today().strftime("%H")<str(18) and datetime.today().strftime("%H")>str(12): #맑은날 점심
                Clear_R = random.choice(Clear_R)
                food = ws[Clear_R].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Clear'and datetime.today().strftime("%H")<str(24) and datetime.today().strftime("%H")>str(18): #맑은날 저녁
                food = ws[Clear_D].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            
            if w.get_status() == 'Clouds'and datetime.today().strftime("%H")<str(12):   #흐린날 아침
                Clouds_B = random.choice(Clouds_B)
                food = ws[Clouds_B].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Clouds'and datetime.today().strftime("%H")<str(18) and datetime.today().strftime("%H")>str(12): #흐린날 점심
                Clouds_R = random.choice(Clouds_R)
                food = ws[Clouds_R].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Clouds'and datetime.today().strftime("%H")<str(24) and datetime.today().strftime("%H")>str(18): #흐린날 저녁
                Clouds_D = random.choice(Clouds_D)
                food = ws[Clouds_D].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)

            if w.get_status() == 'Mist' and datetime.today().strftime("%H")<str(12): #안개낀날 아침
                Mist_B = random.choice(Mist_B)
                food = ws[Mist_B].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Mist' and datetime.today().strftime("%H")<str(18) and datetime.today().strftime("%H")>str(12): #안개낀날 점심
                Mist_R = random.choice(Mist_R)
                food = ws[Mist_R].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Mist' and datetime.today().strftime("%H")<str(24) and datetime.today().strftime("%H")>str(18): #안개낀날 저녁
                Mist_D = random.choice(Mist_D)
                food = ws[Mist_D].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)

            if w.get_status() == 'Haze' and datetime.today().strftime("%H")<str(12): #안개낀날 아침
                Haze_B = random.choice(Haze_B)
                food = ws[Haze_B].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Haze' and datetime.today().strftime("%H")<str(18) and datetime.today().strftime("%H")>str(12): #안개낀날 점심
                Haze_R = random.choice(Haze_R)
                food = ws[Haze_R].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Haze' and datetime.today().strftime("%H")<str(24) and datetime.today().strftime("%H")>str(18): #안개낀날 저녁
                Haze_D = random.choice(Haze_D)
                food = ws[Haze_D].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)

            if w.get_status() == 'Fog' and datetime.today().strftime("%H")<str(12): #안개낀날 아침
                Fog_B = random.choice(Fog_B)
                food = ws[Fog_B].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Fog' and datetime.today().strftime("%H")<str(18) and datetime.today().strftime("%H")>str(12): #안개낀날 점심
                Fog_R = random.choice(Fog_R)
                food = ws[Fog_R].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
            if w.get_status() == 'Fog' and datetime.today().strftime("%H")<str(24) and datetime.today().strftime("%H")>str(18): #안개낀날 저녁
                Fog_D = random.choice(Fog_D)
                food = ws[Fog_D].value
                self.image.setText(food)
                pixmap = QPixmap('아이콘.png')
                self.label.setPixmap(pixmap)
                
            global search #search를 전역변수로 선언하고 search에 음식이름을 넣어줌, 그후 음식 이름 별로 분류
            search = food
            if search == "도넛" or search == "바게뜨" or search == "토스트" or search == "베이글" or search == "모닝빵" or search == "샌드위치" or search == "머핀" or search == "파이":
                search = "베이커리"
            if search == "동지죽" or search == "호박죽":
                search = "죽"
            if search == "유부초밥" or search == "롤" or search == "초밥" :
                search = "초밥"
            if search == "떡튀순" or search == "어묵":
                search = "분식"
            if search == "돼지주물럭" or search == "제육볶음" or search == "오삼불고기":
                search = "백반"
            if search == "돌솥비빔밥":
                search = "비빔밥"
            if search == "닭강정" or search == "치킨&맥주":
                search = "치킨"
            if search == "짬뽕" or search == "짜장" or search == "칠리새우" or search == "깐풍기" or search == "쟁반짜장" or search == "탕수육":
                search = "중국집"
            if search == "갈매기살" or search == "삼겹살":
                search = "고기"
            if search == "쭈꾸미철판볶음":
                search = "쭈꾸미"
            if search == "낚지볶음":
                search = "낚지"
            if search == "족발&보쌈":
                search = "족발"
            if search == "고르곤졸라":
                search = "레스토랑"
            if search == "튀김&맥주":
                search = "호프"
            if search == "김치찌개" or search == "된장찌개":
                search = "찌개"
            if search == "삶은달걀"or search == "오꼬노미야끼"or search == "과일" or search == "붕어빵" or search == "호떡" or search == "타코야끼" or search == "와플" or search == "핫도그" or search == "고구마" or search == "호빵" or search == "미역국" or search == "떡국" or search == "스프" or search == "약밥" or search == "씨리얼" or search == "에그스크램블&베이컨" or search == "주먹밥" or search == "핫케익" or search == "견과류" or search == "과일주스" or search == "오믈렛" or search == "간장계란밥" or search == "샐러드" or search == "볶음밥":
                search = "마트"
            if search == "홍합탕":
                search = "홍합"
            if search == "막걸리&전":
                search = "막걸리"
                
        except: # 오류 발생시 처리 
            print('순서 오류')
            self.image.setText("지역을 입력해 주세요.")
            pixmap = QPixmap('Question.png')
            self.label.setPixmap(pixmap)
            
        
      
        
    def btn2_clicked(self):
        
        try :
            url = "https://www.google.co.kr/maps/search/" + search
            webbrowser.open(url)

        except: #오류 발생시 처리
            print('search 오류')
            self.image.move(35, 200) # QLabel 위치 설정
            self.image.resize(240, 100) # QLabel 크기 설정
        
            self.label.move(75, 70) # QLabel 위치 설정
            self.label.resize(150, 150) # QLabel 크기 설정

            self.image.setText('추천메뉴를 먼저 검색해주세요.')
            pixmap = QPixmap('Question.png')
            self.label.setPixmap(pixmap)
        
    def lineEditPressed(self): #lineEditPressed 선언
        self.label.move(60, 20) # QLabel 위치 설정
        self.label.resize(180, 200) # QLabel 크기 설정
        
        self.image.resize(100,100) # QLabel 크기 설정
        self.image.move(105, 170) # QLabel 위치 설정

        try :
            city = str(self.lineEdit.text()) # city에 lineEdit에 써진 텍스트 입력 
            obs = owm.weather_at_place(city) #날씨를 검색할 장소를 city로 설정
            w = obs.get_weather() # 날씨 정보 가져오기
            l = obs.get_location() # 위치 정보 가져오기
            self.label.setText(l.get_name()+'\n최고 기온 : ' + str(w.get_temperature(unit='celsius')['temp_max']) +
                           '˚C' + '\n최저 기온 : '+ str(w.get_temperature(unit='celsius')['temp_min']) + '˚C'
                           + '\n현재 기온 : ' + str(w.get_temperature(unit='celsius')['temp']) + '˚C' +
                            '\n현재 날씨 : ' + w.get_status()) # QLabel에 날씨정보 출력
                                                        
            if w.get_status() == 'Snow': # 눈오는 날씨면 눈 그림 출력
                pixmap = QPixmap('Snow.png')
                self.image.setPixmap(pixmap)
            
            if w.get_status() == 'Rain': # 비오는 날씨면 비 그림 출력
                pixmap = QPixmap('Rain.png')
                self.image.setPixmap(pixmap)
        
            if w.get_status() == 'Clear': # 화창한 날씨면 해 그림 출력
                pixmap = QPixmap('Clear.png')
                self.image.setPixmap(pixmap)
            
            if w.get_status() == 'Clouds': # 구름낀 날씨면 구름 그림 출력
                pixmap = QPixmap('Clouds.png')
                self.image.setPixmap(pixmap)
            
            if w.get_status() == 'Mist': # 안개낀 날씨면 안개 그림 출력
                pixmap = QPixmap('Mist.png')
                self.image.setPixmap(pixmap)

            if w.get_status() == 'Haze': # 안개낀 날씨면 안개 그림 출력
                pixmap = QPixmap('Mist.png')
                self.image.setPixmap(pixmap)

            if w.get_status() == 'Fog': # 안개낀 날씨면 안개 그림 출력
                pixmap = QPixmap('Mist.png')
                self.image.setPixmap(pixmap)
            
        except: #오류 발생시 처리
            print('지역명 오류')
            self.image.move(45, 200) # QLabel 위치 설정
            self.image.resize(220, 100) # QLabel 크기 설정
        
            self.label.move(75, 70) # QLabel 위치 설정
            self.label.resize(150, 150) # QLabel 크기 설정

            self.image.setText('지역명이 틀렸습니다.')
            pixmap = QPixmap('Question.png')
            self.label.setPixmap(pixmap)
            
            
if __name__ == "__main__": # 창 출력
   
    app = QApplication(sys.argv)
    mywindow = MyWindow()
    mywindow.show()
    app.exec_()

